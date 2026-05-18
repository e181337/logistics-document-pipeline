import io
import re
from datetime import date, datetime, time, timezone
from pathlib import PurePosixPath
from typing import Any

from PIL import Image, UnidentifiedImageError

from app.repositories import DocumentRepository
from app.services.storage import StorageService, parse_gcs_uri


EXCEL_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}
MAX_ROWS_PER_SHEET = 500
MAX_COLUMNS_PER_SHEET = 50
PARSED_TEXT_PREVIEW_LENGTH = 4000


class ExcelService:
    def __init__(self) -> None:
        self.repository = DocumentRepository()
        self.storage = StorageService()

    def parse_workbook(
        self,
        document_id: str,
        tenant_id: str,
        trace_id: str,
        file_uri: str,
    ) -> dict[str, Any]:
        from openpyxl import load_workbook

        workbook_bytes = self.storage.download_bytes(file_uri)
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        parsed_at = datetime.now(timezone.utc)
        parsed_sections: list[str] = []
        sheets: list[dict[str, Any]] = []
        embedded_images: list[dict[str, Any]] = []

        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_id = f"sheet_{sheet_index:04d}"
            sheet_text, row_count, column_count = worksheet_to_text(worksheet)
            sheet_payload = {
                "document_id": document_id,
                "tenant_id": tenant_id,
                "trace_id": trace_id,
                "sheet_id": sheet_id,
                "sheet_name": worksheet.title,
                "sheet_index": sheet_index,
                "row_count": row_count,
                "column_count": column_count,
                "text_preview": sheet_text[:PARSED_TEXT_PREVIEW_LENGTH],
                "status": "PARSED",
                "updated_at": parsed_at,
            }
            self.repository.create_sheet(document_id, sheet_id, sheet_payload)
            sheets.append(
                {
                    "sheet_id": sheet_id,
                    "sheet_name": worksheet.title,
                    "sheet_index": sheet_index,
                    "row_count": row_count,
                    "column_count": column_count,
                }
            )
            if sheet_text:
                parsed_sections.append(f"--- Sheet: {worksheet.title} ---\n{sheet_text}")

            embedded_images.extend(
                self.extract_embedded_images(
                    document_id=document_id,
                    tenant_id=tenant_id,
                    trace_id=trace_id,
                    file_uri=file_uri,
                    sheet_id=sheet_id,
                    sheet_name=worksheet.title,
                    sheet_index=sheet_index,
                    worksheet=worksheet,
                    parsed_at=parsed_at,
                )
            )

        parsed_text = "\n\n".join(parsed_sections)
        parsed_text_uri = self.upload_parsed_text(file_uri, parsed_text)
        return {
            "document_kind": "excel",
            "parser": "openpyxl_v1",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "embedded_image_count": len(embedded_images),
            "embedded_images": embedded_images,
            "parsed_text_uri": parsed_text_uri,
            "parsed_text_preview": parsed_text[:PARSED_TEXT_PREVIEW_LENGTH],
            "parsed_text_length": len(parsed_text),
            "parsed_at": parsed_at,
        }

    def extract_embedded_images(
        self,
        document_id: str,
        tenant_id: str,
        trace_id: str,
        file_uri: str,
        sheet_id: str,
        sheet_name: str,
        sheet_index: int,
        worksheet: Any,
        parsed_at: datetime,
    ) -> list[dict[str, Any]]:
        images = getattr(worksheet, "_images", [])
        image_summaries = []
        for image_index, workbook_image in enumerate(images, start=1):
            image_id = f"{sheet_id}_image_{image_index:04d}"
            image_bytes = workbook_image._data()
            content_type, extension, width, height = inspect_image_bytes(image_bytes)
            image_uri = self.upload_embedded_image(
                file_uri=file_uri,
                sheet_name=sheet_name,
                image_id=image_id,
                extension=extension,
                image_bytes=image_bytes,
                content_type=content_type,
            )
            anchor = image_anchor(workbook_image)
            payload = {
                "document_id": document_id,
                "tenant_id": tenant_id,
                "trace_id": trace_id,
                "image_id": image_id,
                "sheet_id": sheet_id,
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "image_index": image_index,
                "image_uri": image_uri,
                "content_type": content_type,
                "width": width,
                "height": height,
                "anchor": anchor,
                "status": "EXTRACTED",
                "updated_at": parsed_at,
            }
            self.repository.create_embedded_image(document_id, image_id, payload)
            image_summaries.append(
                {
                    "image_id": image_id,
                    "sheet_id": sheet_id,
                    "sheet_name": sheet_name,
                    "image_uri": image_uri,
                    "content_type": content_type,
                    "width": width,
                    "height": height,
                    "anchor": anchor,
                }
            )
        return image_summaries

    def upload_parsed_text(self, file_uri: str, parsed_text: str) -> str:
        bucket_name, object_name = parse_gcs_uri(file_uri)
        parsed_object_name = f"{parent_object_name(object_name)}/excel/parsed_workbook.txt"
        stored_file = self.storage.upload(
            bucket_name=bucket_name,
            object_name=parsed_object_name,
            content=parsed_text.encode("utf-8"),
            content_type="text/plain; charset=utf-8",
        )
        return stored_file.file_uri

    def upload_embedded_image(
        self,
        file_uri: str,
        sheet_name: str,
        image_id: str,
        extension: str,
        image_bytes: bytes,
        content_type: str,
    ) -> str:
        bucket_name, object_name = parse_gcs_uri(file_uri)
        safe_sheet_name = safe_object_segment(sheet_name)
        image_object_name = (
            f"{parent_object_name(object_name)}/excel/images/"
            f"{safe_sheet_name}/{image_id}.{extension}"
        )
        stored_file = self.storage.upload(
            bucket_name=bucket_name,
            object_name=image_object_name,
            content=image_bytes,
            content_type=content_type,
        )
        return stored_file.file_uri


def worksheet_to_text(worksheet: Any) -> tuple[str, int, int]:
    lines = []
    row_count = 0
    column_count = 0
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(worksheet.max_row or 1, MAX_ROWS_PER_SHEET),
        max_col=min(worksheet.max_column or 1, MAX_COLUMNS_PER_SHEET),
        values_only=True,
    ):
        values = [format_cell_value(value) for value in row]
        trimmed_values = trim_trailing_empty(values)
        if not trimmed_values:
            continue
        row_count += 1
        column_count = max(column_count, len(trimmed_values))
        lines.append("\t".join(trimmed_values))
    return "\n".join(lines), row_count, column_count


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value).strip()


def trim_trailing_empty(values: list[str]) -> list[str]:
    trimmed = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def inspect_image_bytes(image_bytes: bytes) -> tuple[str, str, int | None, int | None]:
    try:
        with Image.open(io.BytesIO(image_bytes)) as image:
            image_format = (image.format or "PNG").lower()
            extension = "jpg" if image_format == "jpeg" else image_format
            content_type = "image/jpeg" if extension == "jpg" else f"image/{extension}"
            return content_type, extension, image.width, image.height
    except UnidentifiedImageError:
        return "application/octet-stream", "bin", None, None


def image_anchor(workbook_image: Any) -> dict[str, int] | None:
    anchor = getattr(workbook_image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return {
        "row": marker.row + 1,
        "column": marker.col + 1,
    }


def parent_object_name(object_name: str) -> str:
    return str(PurePosixPath(object_name).parent)


def safe_object_segment(value: str) -> str:
    safe_value = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return safe_value or "sheet"
